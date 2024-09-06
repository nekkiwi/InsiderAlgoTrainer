import os
import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.svm import SVC
from sklearn.gaussian_process import GaussianProcessClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.metrics import matthews_corrcoef, f1_score, accuracy_score, confusion_matrix, mean_squared_error
from sklearn.model_selection import cross_val_predict, StratifiedKFold
from tqdm import tqdm
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import joblib  # To save models

def get_model(model_type):
    """Return the model instance based on the specified model type."""
    if model_type == "RandomForest":
        return RandomForestClassifier(random_state=42)
    elif model_type == "NaivesBayes":
        return GaussianNB()
    elif model_type == "RBF SVM":
        return SVC(kernel='rbf', probability=True, random_state=42)
    elif model_type == "Gaussian Process":
        return GaussianProcessClassifier(random_state=42)
    elif model_type == "Neural Net":
        return MLPClassifier(random_state=42, max_iter=1000)
    else:
        raise ValueError(f"Unknown model type: {model_type}")

def load_data(features_file, targets_file, selected_features_file, target_name):
    """Load the data from Excel files."""
    target_name_clean = target_name.replace(" ", "-").lower()
    features_df = pd.read_excel(features_file)
    targets_df = pd.read_excel(targets_file, sheet_name=None)
    selected_features_df = pd.read_excel(selected_features_file, sheet_name=target_name_clean)
    
    return features_df, targets_df, selected_features_df

def train_model_task(row, features_df, targets_df, target_name, model):
    """Train the model for a specific limit/stop combination and return the results."""
    limit, stop, selected_features_str = row['Limit'], row['Stop'], row['Selected Features']
    selected_features = eval(selected_features_str)
    selected_features = [feature.strip("'") for feature in selected_features]

    # Select the features and target
    X = features_df[selected_features]
    y = targets_df[f'lim {limit} stop {stop}'][target_name]

    # Perform 5-fold cross-validation
    skf = StratifiedKFold(n_splits=5)
    predictions = cross_val_predict(model, X, y, cv=skf)

    if isinstance(y.iloc[0], (int, np.int64, np.int32)):  # Binary target
        # Binary target metrics
        mcc = matthews_corrcoef(y, predictions)
        f1_positive = f1_score(y, predictions, pos_label=1)
        f1_negative = f1_score(y, predictions, pos_label=0)
        acc = accuracy_score(y, predictions)
        cm = confusion_matrix(y, predictions)

        result = {
            "Limit": limit,
            "Stop": stop,
            "Accuracy": acc,
            "MCC": mcc,
            "F1_pos": f1_positive,
            "F1_neg": f1_negative,
            "TN": cm[0, 0],
            "FP": cm[0, 1],
            "FN": cm[1, 0],
            "TP": cm[1, 1],
        }

        prediction_data = pd.DataFrame({
            "Ticker": features_df['Ticker'],
            "Filing Date": features_df['Filing Date'].dt.strftime('%d/%m/%Y %H:%M'),
            f'GT_{target_name}': y,
            f'Pred_{target_name}': predictions
        })

    else:
        # Continuous target metrics
        mse = mean_squared_error(y, predictions)
        y_sign = np.sign(y)
        pred_sign = np.sign(predictions)

        mcc = matthews_corrcoef(y_sign, pred_sign)
        f1_positive = f1_score(y_sign, pred_sign, pos_label=1)
        f1_negative = f1_score(y_sign, pred_sign, pos_label=0)
        acc = accuracy_score(y_sign, pred_sign)
        cm = confusion_matrix(y_sign, pred_sign)

        result = {
            "Limit": limit,
            "Stop": stop,
            "Accuracy": acc,
            "MSE": mse,
            "MCC": mcc,
            "F1_pos": f1_positive,
            "F1_neg": f1_negative,
            "TN": cm[0, 0],
            "FP": cm[0, 1],
            "FN": cm[1, 0],
            "TP": cm[1, 1]
        }

        prediction_data = pd.DataFrame({
            "Ticker": features_df['Ticker'],
            "Filing Date": features_df['Filing Date'].dt.strftime('%d/%m/%Y %H:%M'),
            f'GT_{target_name}': y,
            f'Pred_{target_name}': predictions
        })

    return result, prediction_data

def save_results(results, output_dir, model_type, target_name):
    """Save the results to an Excel file, creating a new subsheet if necessary."""
    stats_file = os.path.join(output_dir, f'stats_{model_type.replace(" ", "-").lower()}.xlsx')
    os.makedirs(os.path.dirname(stats_file), exist_ok=True)

    results_df = pd.DataFrame(results)

    with pd.ExcelWriter(stats_file, engine='openpyxl', mode='a' if os.path.exists(stats_file) else 'w') as writer:
        if target_name in writer.book.sheetnames:
            del writer.book[target_name]
        results_df.to_excel(writer, sheet_name=target_name, index=False)

def save_predictions(all_predictions, predictions_dir, model_short):
    """Save predictions to individual Excel files."""
    model_subdir = os.path.join(predictions_dir, model_short)
    os.makedirs(model_subdir, exist_ok=True)

    pbar = tqdm(total=len(all_predictions), desc="Saving Predictions")

    for pred_data in all_predictions:
        limit = pred_data["Limit"]
        stop = pred_data["Stop"]
        df = pred_data["Predictions"]
        prediction_file = os.path.join(model_subdir, f'pred_{model_short}_l{limit}_s{stop}.xlsx')

        if os.path.exists(prediction_file):
            book = load_workbook(prediction_file)
        else:
            book = Workbook()
            book.remove(book.active)

        sheet_name = f'pred_{model_short}_l{limit}_s{stop}'
        if sheet_name in book.sheetnames:
            del book[sheet_name]

        sheet = book.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            sheet.append(r)

        book.save(prediction_file)
        pbar.update(1)

    pbar.close()

def save_model(model, model_name, target_name, criterion, output_dir):
    """Save the model to a file."""
    model_dir = os.path.join(output_dir, f'{model_name}_{target_name}_{criterion}')
    os.makedirs(os.path.dirname(model_dir), exist_ok=True)
    joblib.dump(model, model_dir)
