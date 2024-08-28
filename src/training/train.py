import os
import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.svm import SVC
from sklearn.gaussian_process import GaussianProcessClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.model_selection import cross_val_predict, StratifiedKFold
from sklearn.metrics import confusion_matrix, f1_score, matthews_corrcoef, mean_squared_error, accuracy_score
import sys
from multiprocessing import Pool, cpu_count
from tqdm import tqdm
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Add the 'src' directory to the system path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

class ModelTrainer:
    def __init__(self):
        data_dir = os.path.join(os.path.dirname(__file__), '../../data')
        self.features_file = os.path.join(data_dir, "final/features_final.xlsx")
        self.targets_file = os.path.join(data_dir, "final/targets_final.xlsx")
        self.selected_features_file = os.path.join(data_dir, "output/feature_selection/selected_features.xlsx")
        self.output_dir = os.path.join(data_dir, "training/stats")
        self.predictions_dir = os.path.join(data_dir, "training/predictions")
        self.target_name = ""
        self.model_type = "RandomForest"  # Default model

    def load_data(self, target_name):
        """Load the data from the Excel files."""
        self.target_name = target_name.replace(" ", "-").lower()
        self.features_df = pd.read_excel(self.features_file)
        self.targets_df = pd.read_excel(self.targets_file, sheet_name=None)
        self.selected_features_df = pd.read_excel(self.selected_features_file, sheet_name=self.target_name)

    def get_model(self):
        """Return the model instance based on the specified model type."""
        if self.model_type == "RandomForest":
            return RandomForestClassifier(random_state=42)
        elif self.model_type == "NaivesBayes":
            return GaussianNB()
        elif self.model_type == "RBF SVM":
            return SVC(kernel='rbf', probability=True, random_state=42)
        elif self.model_type == "Gaussian Process":
            return GaussianProcessClassifier(random_state=42)
        elif self.model_type == "Neural Net":
            return MLPClassifier(random_state=42, max_iter=1000)
        else:
            raise ValueError(f"Unknown model type: {self.model_type}")

    def train_model_task(self, row):
        """Train the model for a specific limit/stop combination and return the results."""
        limit, stop, selected_features_str = row['Limit'], row['Stop'], row['Selected Features']
        selected_features = eval(selected_features_str)
        selected_features = [feature.strip("'") for feature in selected_features]

        # Select the features and target
        X = self.features_df[selected_features]
        y = self.targets_df[f'lim {limit} stop {stop}'][self.target_name]

        # Initialize the model
        model = self.get_model()

        # Perform 5-fold cross-validation
        skf = StratifiedKFold(n_splits=5)
        predictions = cross_val_predict(model, X, y, cv=skf)

        prediction_data = None

        if self.binary_target:
            # Calculate metrics for binary targets
            mcc = matthews_corrcoef(y, predictions)
            f1_positive = f1_score(y, predictions, pos_label=1)
            f1_negative = f1_score(y, predictions, pos_label=0)
            acc = accuracy_score(y, predictions)
            cm = confusion_matrix(y, predictions)

            result = {
                "Limit": limit,
                "Stop": stop,
                "Accuracy": acc,
                "MCC": mcc,
                "F1_pos": f1_positive,
                "F1_neg": f1_negative,
                "TN": cm[0, 0],
                "FP": cm[0, 1],
                "FN": cm[1, 0],
                "TP": cm[1, 1],
            }

            prediction_data = {
                "sheet_name": f'pred_{self.model_type.replace(" ", "-").lower()}_l{limit}_s{stop}',
                "limit": limit,
                "stop": stop,
                "data": pd.DataFrame({
                    "Ticker": self.features_df['Ticker'],
                    "Filing Date": self.features_df['Filing Date'],
                    f'GT_{self.target_name}': y,
                    f'Pred_{self.target_name}': predictions
                })
            }

        else:
            # Calculate metrics for continuous targets
            mse = mean_squared_error(y, predictions)
            y_sign = np.sign(y)
            pred_sign = np.sign(predictions)

            mcc = matthews_corrcoef(y_sign, pred_sign)
            f1_positive = f1_score(y_sign, pred_sign, pos_label=1)
            f1_negative = f1_score(y_sign, pred_sign, pos_label=0)
            acc = accuracy_score(y_sign, pred_sign)
            cm = confusion_matrix(y_sign, pred_sign)

            result = {
                "Limit": limit,
                "Stop": stop,
                "Accuracy": acc,
                "MSE": mse,
                "MCC": mcc,
                "F1_pos": f1_positive,
                "F1_neg": f1_negative,
                "TN": cm[0, 0],
                "FP": cm[0, 1],
                "FN": cm[1, 0],
                "TP": cm[1, 1]
            }

            prediction_data = {
                "sheet_name": f'pred_{self.model_type.replace(" ", "-").lower()}_l{limit}_s{stop}',
                "limit": limit,
                "stop": stop,
                "data": pd.DataFrame({
                    "Ticker": self.features_df['Ticker'],
                    "Filing Date": self.features_df['Filing Date'],
                    f'GT_{self.target_name}': y,
                    f'Pred_{self.target_name}': predictions
                })
            }

        return result, prediction_data


    def train_model(self):
        """Train the model for each limit/stop combination in parallel using multiprocessing.Pool."""
        rows = self.selected_features_df.to_dict(orient='records')
        all_predictions = []

        # Use multiprocessing Pool to parallelize the training process
        with Pool(cpu_count()) as pool:
            results = list(tqdm(pool.imap(self.train_model_task, rows), total=len(rows), desc="Training Models"))

        # Filter out None results and collect predictions
        filtered_results = []
        for res, pred_data in results:
            if res is not None:
                filtered_results.append(res)
            if pred_data is not None:
                all_predictions.append(pred_data)

        # Save the results if any
        if filtered_results:
            self.save_results(filtered_results)

        # Save all predictions to the Excel file at once
        self.save_predictions(all_predictions)

    def save_results(self, results):
        """Save the results to the Excel file, creating a new subsheet if necessary."""
        stats_file = os.path.join(self.output_dir, f'stats_{self.model_type.replace(" ", "-").lower()}.xlsx')
        os.makedirs(os.path.dirname(stats_file), exist_ok=True)

        results_df = pd.DataFrame(results)

        with pd.ExcelWriter(stats_file, engine='openpyxl', mode='a' if os.path.exists(stats_file) else 'w') as writer:
            if self.target_name.replace(" ", "-").lower() in writer.book.sheetnames:
                # If the sheet exists, remove it first
                del writer.book[self.target_name.replace(" ", "-").lower()]
            results_df.to_excel(writer, sheet_name=self.target_name.replace(" ", "-").lower(), index=False)
        print(f"Results saved to {stats_file}")

    def save_predictions(self, all_predictions):
        """Save all predictions stored in memory to individual Excel files."""
        # Create a subdirectory for the model under the predictions directory
        model_subdir = os.path.join(self.predictions_dir, self.model_type.replace(" ", "-").lower())
        os.makedirs(model_subdir, exist_ok=True)

        # Initialize the progress bar
        pbar = tqdm(total=len(all_predictions), desc="Saving Predictions")

        # Iterate over all prediction data
        for pred_data in all_predictions:
            limit = pred_data["limit"]
            stop = pred_data["stop"]
            df = pred_data["data"]

            if self.model_type.replace(" ", "-").lower() == 'randomforest': model_short = 'rf'
            elif self.model_type.replace(" ", "-").lower() == 'naivesbayes': model_short = 'nb'
            elif self.model_type.replace(" ", "-").lower() == 'rbf-svm': model_short = 'svm'
            elif self.model_type.replace(" ", "-").lower() == 'gaussian-process': model_short = 'gp'
            elif self.model_type.replace(" ", "-").lower() == 'neural-net': model_short = 'nn'

            # File name based on model type, limit, and stop values
            prediction_file = os.path.join(model_subdir, f'pred_{model_short}_l{limit}_s{stop}.xlsx')

            if os.path.exists(prediction_file):
                # Load existing workbook
                book = load_workbook(prediction_file)
            else:
                # Create a new workbook if it doesn't exist
                book = Workbook()
                book.remove(book.active)  # Remove the default sheet created with the new workbook

            # Since we now have one sheet per file, we can use a consistent sheet name or just the default sheet
            sheet_name = f'pred_{model_short}_l{limit}_s{stop}'
            
            if sheet_name in book.sheetnames:
                # If the sheet already exists, update it
                sheet = book[sheet_name]
                
                # Convert the existing sheet to a DataFrame
                existing_df = pd.DataFrame(sheet.values)
                existing_df.columns = existing_df.iloc[0]
                existing_df = existing_df[1:]

                # Update or append the columns
                for col in df.columns:
                    if col in existing_df.columns:
                        existing_df[col] = df[col]
                    else:
                        existing_df = pd.concat([existing_df, df[[col]]], axis=1)

                # Remove the existing sheet
                del book[sheet_name]
                
                # Add the updated sheet back to the workbook
                sheet = book.create_sheet(title=sheet_name)
                for r in dataframe_to_rows(existing_df, index=False, header=True):
                    sheet.append(r)

            else:
                # If the sheet doesn't exist, create a new one
                sheet = book.create_sheet(title=sheet_name)
                for r in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(r)

            # Save the workbook
            book.save(prediction_file)
            
            # print(f"Predictions saved to {prediction_file}")
            
            # Update the progress bar
            pbar.update(1)

        # Close the progress bar
        pbar.close()
        print(f"Predictions saved to {self.predictions_dir}")

    def run(self, target_name, model_type="RandomForest"):
        """Run the full training process."""
        self.binary_target = target_name not in ['Return at cashout', 'Days at cashout']
        self.model_type = model_type  # Set the model type
        self.load_data(target_name)
        self.train_model()

if __name__ == "__main__":
    trainer = ModelTrainer()
    trainer.run(target_name="Limit occurred first", model_type="RandomForest")
